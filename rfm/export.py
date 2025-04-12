import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def export_segments_to_excel(rfm: pd.DataFrame, summary: pd.DataFrame, path: str = "output/rfm_report.xlsx"):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Customer RFM"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(fill_type="solid", fgColor="1B3A6B")
    for r in dataframe_to_rows(rfm[["customer_id","recency","frequency","monetary","segment","clv"]]
                               .round(2), index=False, header=True):
        ws1.append(r)
    for cell in ws1[1]:
        cell.font = hf; cell.fill = hfill
    ws2 = wb.create_sheet("Segment Summary")
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws2.append(r)
    for cell in ws2[1]:
        cell.font = hf; cell.fill = hfill
    wb.save(path)
    print(f"Saved: {path}")
